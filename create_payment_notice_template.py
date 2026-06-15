from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.page import PageMargins


BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "document_templates"
TEMPLATE_PATH = TEMPLATE_DIR / "payment_notice_template.xlsx"


TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

LABEL_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

NOTICE_FILL = PatternFill(
    fill_type="solid",
    fgColor="F3F6F9",
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def apply_range_style(
    worksheet,
    cell_range: str,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> None:
    """指定範囲内のすべてのセルへ書式を設定する。"""

    min_col, min_row, max_col, max_row = range_boundaries(
        cell_range
    )

    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            if font is not None:
                cell.font = font

            if fill is not None:
                cell.fill = fill

            if alignment is not None:
                cell.alignment = alignment

            if border is not None:
                cell.border = border


def merge_and_set(
    worksheet,
    cell_range: str,
    value: str = "",
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> None:
    """セルを結合し、値と書式を設定する。"""

    worksheet.merge_cells(cell_range)

    start_cell = cell_range.split(":")[0]
    worksheet[start_cell] = value

    apply_range_style(
        worksheet,
        cell_range,
        font=font,
        fill=fill,
        alignment=alignment,
        border=border,
    )


def create_template() -> None:
    """仕様v1に対応した支払通知書を作成する。"""

    TEMPLATE_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "支払通知書"

    worksheet.sheet_view.showGridLines = False

    # タイトル
    merge_and_set(
        worksheet,
        "A1:F1",
        "支払通知書",
        font=Font(
            size=20,
            bold=True,
            color="FFFFFF",
        ),
        fill=TITLE_FILL,
        alignment=Alignment(
            horizontal="center",
            vertical="center",
        ),
        border=THIN_BORDER,
    )

    # ラベルと値欄
    field_rows = [
        ("A3:B3", "発行日", "C3:F3"),
        ("A4:B4", "発行番号", "C4:F4"),
        ("A6:B6", "会社名", "C6:F6"),
        ("A7:B7", "住所", "C7:F7"),
        ("A8:B8", "事業者登録番号", "C8:F8"),
        ("A13:B13", "支払金額", "C13:F13"),
        ("A14:B14", "支払日", "C14:F14"),
        ("A15:B15", "振込先", "C15:F15"),
        ("A17:B19", "内訳", "C17:F19"),
        ("A20:B20", "税区分", "C20:F20"),
        ("A22:B24", "備考", "C22:F24"),
    ]

    for label_range, label, value_range in field_rows:
        merge_and_set(
            worksheet,
            label_range,
            label,
            font=Font(bold=True),
            fill=LABEL_FILL,
            alignment=Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            ),
            border=THIN_BORDER,
        )

        merge_and_set(
            worksheet,
            value_range,
            "",
            alignment=Alignment(
                vertical="center",
                wrap_text=True,
            ),
            border=THIN_BORDER,
        )

    # 通知文
    merge_and_set(
        worksheet,
        "A10:F11",
        (
            "下記のとおりお支払い予定をご通知いたします。\n"
            "内容をご確認くださいますようお願いいたします。"
        ),
        fill=NOTICE_FILL,
        alignment=Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        ),
        border=THIN_BORDER,
    )

    # 振込先は固定文言
    worksheet["C15"] = "ご登録振込先"

    # 注意事項
    merge_and_set(
        worksheet,
        "A26:F27",
        (
            "本書は開発用のサンプル帳票です。\n"
            "実際の支払内容については、発行元へ"
            "お問い合わせください。"
        ),
        font=Font(
            size=9,
            color="666666",
        ),
        fill=NOTICE_FILL,
        alignment=Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        ),
        border=THIN_BORDER,
    )

    # 発行元情報
    merge_and_set(
        worksheet,
        "D29:F29",
        "発行元：株式会社サンプル",
        font=Font(bold=True),
        alignment=Alignment(
            horizontal="left",
            vertical="center",
        ),
    )

    merge_and_set(
        worksheet,
        "D30:F30",
        "所在地：福岡県福岡市〇〇区〇〇",
        alignment=Alignment(
            horizontal="left",
            vertical="center",
        ),
    )

    merge_and_set(
        worksheet,
        "D31:F31",
        "問い合わせ先：sample@example.com",
        alignment=Alignment(
            horizontal="left",
            vertical="center",
        ),
    )

    # 各入力欄の文字配置
    apply_range_style(
        worksheet,
        "C3:F3",
        alignment=Alignment(
            horizontal="left",
            vertical="center",
        ),
    )

    apply_range_style(
        worksheet,
        "C4:F4",
        alignment=Alignment(
            horizontal="left",
            vertical="center",
        ),
    )

    apply_range_style(
        worksheet,
        "C6:F6",
        font=Font(
            size=12,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="left",
            vertical="center",
            shrink_to_fit=True,
        ),
    )

    apply_range_style(
        worksheet,
        "C7:F7",
        alignment=Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        ),
    )

    apply_range_style(
        worksheet,
        "C8:F8",
        alignment=Alignment(
            horizontal="left",
            vertical="center",
        ),
    )

    apply_range_style(
        worksheet,
        "C13:F13",
        font=Font(
            size=14,
            bold=True,
        ),
        alignment=Alignment(
            horizontal="right",
            vertical="center",
        ),
    )

    apply_range_style(
        worksheet,
        "C14:F14",
        alignment=Alignment(
            horizontal="left",
            vertical="center",
        ),
    )

    apply_range_style(
        worksheet,
        "C15:F15",
        alignment=Alignment(
            horizontal="left",
            vertical="center",
        ),
    )

    apply_range_style(
        worksheet,
        "C17:F19",
        alignment=Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        ),
    )

    apply_range_style(
        worksheet,
        "C20:F20",
        alignment=Alignment(
            horizontal="left",
            vertical="center",
        ),
    )

    apply_range_style(
        worksheet,
        "C22:F24",
        alignment=Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        ),
    )

    # 値欄の表示形式
    worksheet["C3"].number_format = "yyyy年m月d日"
    worksheet["C13"].number_format = '#,##0"円"'
    worksheet["C14"].number_format = "yyyy年m月d日"

    # 列幅
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 10
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 18
    worksheet.column_dimensions["F"].width = 18

    # 行の高さ
    worksheet.row_dimensions[1].height = 34
    worksheet.row_dimensions[3].height = 24
    worksheet.row_dimensions[4].height = 24
    worksheet.row_dimensions[6].height = 34
    worksheet.row_dimensions[7].height = 46
    worksheet.row_dimensions[8].height = 28
    worksheet.row_dimensions[10].height = 28
    worksheet.row_dimensions[11].height = 28
    worksheet.row_dimensions[13].height = 28
    worksheet.row_dimensions[14].height = 28
    worksheet.row_dimensions[15].height = 28
    worksheet.row_dimensions[17].height = 30
    worksheet.row_dimensions[18].height = 30
    worksheet.row_dimensions[19].height = 30
    worksheet.row_dimensions[20].height = 28
    worksheet.row_dimensions[22].height = 30
    worksheet.row_dimensions[23].height = 30
    worksheet.row_dimensions[24].height = 30
    worksheet.row_dimensions[26].height = 24
    worksheet.row_dimensions[27].height = 24

    # 印刷設定
    worksheet.print_area = "A1:F32"
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.paperSize = (
        worksheet.PAPERSIZE_A4
    )
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.page_margins = PageMargins(
        left=0.4,
        right=0.4,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2,
    )

    workbook.save(TEMPLATE_PATH)

    print(
        "仕様v1対応テンプレートを作成しました："
        f"{TEMPLATE_PATH}"
    )


if __name__ == "__main__":
    create_template()